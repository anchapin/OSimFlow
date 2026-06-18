#!/usr/bin/env python3
"""excel_to_variables.py — PAT/Analysis Gem Excel spreadsheet to variables.yml converter.

Reads a PAT-style ``.xlsx`` spreadsheet and produces a OSimFlow
``variables.yml`` file. Supports the standard PAT column layout:

================  ==============================================================
Column           Description
================  ==============================================================
var_name         Variable name (required)
lower_bound      Lower bound for continuous distributions
upper_bound      Upper bound for continuous distributions
distribution     Distribution name: uniform, normal, lognormal, triangular,
                 discrete, categorical
mean             Mean (normal, lognormal)
stddev           Standard deviation (normal, lognormal)
mode             Mode / peak (triangular)
values           Comma-separated list of values (discrete, categorical)
pmf              Comma-separated probability weights for discrete variables
                 (qdiscrete weighted sampling; issue #579). When omitted,
                 all values have equal probability.
display_name     Optional human-readable label
measure_argument Optional MeasureName.argument_name dotted reference
================  ==============================================================

Usage::

    python -m osimflow._work_scripts.excel_to_variables \\
      --input my_pat_analysis.xlsx \\
      --output variables.yml

    # With custom sheet name
    python -m osimflow._work_scripts.excel_to_variables \\
      --input my_pat_analysis.xlsx \\
      --output variables.yml \\
      --sheet "CustomVariables"

    # With custom column mapping
    python -m osimflow._work_scripts.excel_to_variables \\
      --input my_pat_analysis.xlsx \\
      --output variables.yml \\
      --col-var-name Name \\
      --col-lower LBound \\
      --col-upper UBound
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path
from typing import Any

import yaml

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("excel_to_variables")

# Supported distribution names (must match variables.yml schema).
_SUPPORTED_DISTRIBUTIONS = frozenset(
    {
        "uniform",
        "normal",
        "lognormal",
        "triangular",
        "discrete",
        "categorical",
        "static",
    }
)

# Default column names in a standard PAT spreadsheet.
_DEFAULT_COLUMN_MAP: dict[str, str] = {
    "name": "var_name",
    "min": "lower_bound",
    "max": "upper_bound",
    "distribution": "distribution",
    "mean": "mean",
    "sigma": "stddev",
    "mode": "mode",
    "values": "values",
    "pmf": "pmf",
    "display_name": "display_name",
    "measure_argument": "measure_argument",
}


def _normalize_dist(dist: str) -> str:
    """Normalise a distribution name to the OSimFlow schema.

    Handles common OSA / PAT aliases:
      - ``normal`` → ``normal``
      - ``lognormal`` / ``lognormal_uncertain`` → ``lognormal``
      - ``discrete`` / ``enum`` → ``discrete``
      - ``categorical`` / ``pivot`` → ``categorical``
      - ``uniform`` → ``uniform``
      - ``triangular`` → ``triangular``
    """
    if not dist:
        return "uniform"
    dist_lower = dist.lower().strip()
    if dist_lower in ("lognormal_uncertain", "lognormal uncertain"):
        return "lognormal"
    if dist_lower in ("enum",):
        return "discrete"
    if dist_lower in ("pivot",):
        return "categorical"
    return dist_lower


def _parse_values(values_str: str | None) -> list[Any] | None:
    """Parse a comma-separated values string into a list.

    Handles both numeric and string values. Strips whitespace.
    Returns None if the input is empty or None.
    """
    if not values_str:
        return None
    values_str = values_str.strip()
    if not values_str:
        return None
    result: list[Any] = []
    for token in values_str.split(","):
        token = token.strip()
        if not token:
            continue
        # Try numeric parse first.
        try:
            result.append(int(token))
        except ValueError:
            try:
                result.append(float(token))
            except ValueError:
                result.append(token)
    return result if result else None


def _row_to_variable(row: dict[str, Any], column_map: dict[str, str]) -> dict[str, Any] | None:
    """Convert a spreadsheet row dict to a variables.yml entry.

    Parameters
    ----------
    row
        A dict mapping internal column names to cell values.
    column_map
        Maps internal keys (name, min, max, ...) to actual spreadsheet
        column headers.

    Returns
    -------
    dict or None
        A variables.yml entry dict, or None if the row is empty/invalid.
    """

    # Resolve column names for this row.
    def col(key: str) -> Any:
        header = column_map.get(key, key)
        return row.get(header)

    name = col("name")
    if not name or not str(name).strip():
        return None
    name = str(name).strip()

    dist_raw = col("distribution")
    dist = _normalize_dist(str(dist_raw) if dist_raw else "uniform")

    entry: dict[str, Any] = {"name": name, "distribution": dist}

    # Attach optional display_name.
    display_name = col("display_name")
    if display_name and str(display_name).strip():
        entry["display_name"] = str(display_name).strip()

    # Attach optional measure_argument.
    measure_arg = col("measure_argument")
    if measure_arg and str(measure_arg).strip():
        entry["measure_argument"] = str(measure_arg).strip()

    # Distribution-specific parameters.
    if dist == "static":
        # Static: just a fixed value (if any).
        static_val = col("value") or col("default_value")
        if static_val is not None:
            entry["value"] = static_val
        return entry

    if dist in ("uniform", "triangular"):
        min_val = col("min")
        max_val = col("max")
        if min_val is not None:
            entry["min"] = float(min_val)
        if max_val is not None:
            entry["max"] = float(max_val)
        if dist == "triangular":
            mode_val = col("mode")
            if mode_val is not None:
                entry["mode"] = float(mode_val)
        return entry

    if dist in ("normal", "lognormal"):
        mean_val = col("mean")
        sigma_val = col("sigma")
        if mean_val is not None:
            entry["mean"] = float(mean_val)
        if sigma_val is not None:
            entry["sigma"] = float(sigma_val)
        return entry

    if dist in ("discrete", "categorical"):
        values_str = col("values")
        values_list = _parse_values(values_str)
        if values_list is not None:
            entry["values"] = values_list
        pmf_str = col("pmf")
        pmf_list = _parse_values(pmf_str)
        if pmf_list is not None:
            pmf_values: list[Any] = values_list if values_list is not None else []
            if len(pmf_list) != len(pmf_values):
                log.warning(
                    "pmf and values lists have different lengths for variable %r "
                    "(%d vs %d); using uniform probabilities",
                    name,
                    len(pmf_list),
                    len(pmf_values),
                )
            else:
                entry["discrete_distribution"] = {
                    "pmf": {str(v): float(p) for v, p in zip(pmf_values, pmf_list, strict=True)},
                }
        return entry

    # Unknown distribution — still include what we have.
    log.warning(
        "Unknown distribution %r for variable %r; including as-is",
        dist,
        name,
    )
    return entry


def excel_to_variables_yml(
    excel_path: Path,
    output_path: Path,
    sheet_name: str = "Variables",
    algorithm: str = "lhs",
    column_map: dict[str, str] | None = None,
) -> None:
    """Convert a PAT Excel spreadsheet to a OSimFlow ``variables.yml`` file.

    Parameters
    ----------
    excel_path
        Path to the ``.xlsx`` file.
    output_path
        Where to write the resulting ``variables.yml``.
    sheet_name
        Name of the worksheet to read. Defaults to ``"Variables"``.
    algorithm
        Sampling / optimisation algorithm. Defaults to ``"lhs"``.
    column_map
        Optional override for the column name mapping. Keys are internal
        identifiers (``name``, ``min``, ``max``, ``distribution``, ``mean``,
        ``sigma``, ``mode``, ``values``, ``display_name``, ``measure_argument``)
        and values are the actual column header strings in the spreadsheet.
        If omitted, the :data:`_DEFAULT_COLUMN_MAP` is used.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read Excel files. Install it with: pip install openpyxl"
        ) from exc

    col_map = dict(_DEFAULT_COLUMN_MAP)
    if column_map:
        col_map.update(column_map)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        available = ", ".join(sorted(wb.sheetnames))
        raise ValueError(
            f"Sheet {sheet_name!r} not found in {excel_path}. Available sheets: {available}"
        )
    ws = wb[sheet_name]

    # Read header row.
    rows_iter = iter(ws.iter_rows(values_only=True))
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError(f"Sheet {sheet_name!r} is empty in {excel_path}")

    # Build a header → column-index map.
    header_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            header_map[str(cell).strip()] = idx

    # Detect which columns are present.
    log.info(
        "Found columns: %s",
        ", ".join(sorted(header_map.keys())),
    )

    variables: list[dict[str, Any]] = []

    for row in rows_iter:
        # Build a dict mapping header name → cell value.
        row_dict: dict[str, Any] = {}
        for header, col_idx in header_map.items():
            if col_idx < len(row):
                row_dict[header] = row[col_idx]

        var_entry = _row_to_variable(row_dict, col_map)
        if var_entry is None:
            # Empty row — skip silently.
            continue
        variables.append(var_entry)

    if not variables:
        raise ValueError(
            f"No valid variables found in sheet {sheet_name!r}. "
            f"Check that the 'var_name' column is present and rows are not empty."
        )

    result: dict[str, Any] = {
        "algorithm": algorithm,
        "variables": variables,
    }

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        yaml.dump(result, f, default_flow_style=False, sort_keys=False)

    log.info(
        "Wrote %d variables to %s (algorithm=%s)",
        len(variables),
        output_path,
        algorithm,
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="Path to the input .xlsx file",
    )
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="Path for the output variables.yml file",
    )
    parser.add_argument(
        "--sheet",
        default="Variables",
        help="Worksheet name to read (default: Variables)",
    )
    parser.add_argument(
        "--algorithm",
        default="lhs",
        help="Algorithm name for variables.yml (default: lhs)",
    )
    # Optional custom column overrides.
    parser.add_argument(
        "--col-var-name",
        dest="col_var_name",
        metavar="HEADER",
        help="Column header for variable name (default: var_name)",
    )
    parser.add_argument(
        "--col-lower",
        dest="col_lower",
        metavar="HEADER",
        help="Column header for lower bound (default: lower_bound)",
    )
    parser.add_argument(
        "--col-upper",
        dest="col_upper",
        metavar="HEADER",
        help="Column header for upper bound (default: upper_bound)",
    )
    parser.add_argument(
        "--col-distribution",
        dest="col_distribution",
        metavar="HEADER",
        help="Column header for distribution (default: distribution)",
    )
    parser.add_argument(
        "--col-mean",
        dest="col_mean",
        metavar="HEADER",
        help="Column header for mean (default: mean)",
    )
    parser.add_argument(
        "--col-stddev",
        dest="col_stddev",
        metavar="HEADER",
        help="Column header for stddev (default: stddev)",
    )
    parser.add_argument(
        "--col-mode",
        dest="col_mode",
        metavar="HEADER",
        help="Column header for mode (default: mode)",
    )
    parser.add_argument(
        "--col-values",
        dest="col_values",
        metavar="HEADER",
        help="Column header for values list (default: values)",
    )
    parser.add_argument(
        "--col-pmf",
        dest="col_pmf",
        metavar="HEADER",
        help="Column header for pmf/weights (default: pmf)",
    )
    parser.add_argument(
        "--col-display-name",
        dest="col_display_name",
        metavar="HEADER",
        help="Column header for display_name (default: display_name)",
    )
    parser.add_argument(
        "--col-measure-argument",
        dest="col_measure_argument",
        metavar="HEADER",
        help="Column header for measure_argument (default: measure_argument)",
    )

    args = parser.parse_args()

    # Build column map from CLI overrides.
    col_map: dict[str, str] = dict(_DEFAULT_COLUMN_MAP)
    if args.col_var_name:
        col_map["name"] = args.col_var_name
    if args.col_lower:
        col_map["min"] = args.col_lower
    if args.col_upper:
        col_map["max"] = args.col_upper
    if args.col_distribution:
        col_map["distribution"] = args.col_distribution
    if args.col_mean:
        col_map["mean"] = args.col_mean
    if args.col_stddev:
        col_map["sigma"] = args.col_stddev
    if args.col_mode:
        col_map["mode"] = args.col_mode
    if args.col_values:
        col_map["values"] = args.col_values
    if args.col_pmf:
        col_map["pmf"] = args.col_pmf
    if args.col_display_name:
        col_map["display_name"] = args.col_display_name
    if args.col_measure_argument:
        col_map["measure_argument"] = args.col_measure_argument

    try:
        excel_to_variables_yml(
            excel_path=args.input,
            output_path=args.output,
            sheet_name=args.sheet,
            algorithm=args.algorithm,
            column_map=col_map,
        )
    except Exception as exc:
        log.error("Conversion failed: %s", exc)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
